from pathlib import Path
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter # Утилита openpyxl для получения буквы столбца таблицы

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.
    Использовать openpyxl ИЛИ xlsxwriter.
    Первая строка CSV — заголовок.
    Лист называется "Sheet1".
    Колонки — автоширина по длине текста (не менее 8 символов).
    """

    csv_path = Path(csv_path)
    xlsx_path = Path(xlsx_path)

    if not csv_path.exists():  # Явная проверка существования файла
        raise FileNotFoundError('Файл не найден')
    
    if not xlsx_path.exists():  # Явная проверка существования файла
        raise FileNotFoundError('Файл не найден')
    
    if csv_path.suffix != '.csv':
            raise ValueError("Неверный тип файла")
    '''
    Работаем с CSV-файлом, читаем из него информацию и проверяем на наличие ошибок. Если они есть - они всплывают.
    '''
    with open(csv_path, 'r', encoding='utf-8') as csv_file:
        reader = csv.DictReader(csv_file)
        csv_file = list(reader) # Сохраняем считанные данные в список
        if len(csv_file) == 0:
             raise ValueError("Пустой файл")
        
        if not reader.fieldnames:
             raise ValueError('Файл не содержит заголовков')
        
        wb = Workbook()      # Создаем рабочую книгу
        ws = wb.active       # Создаём в ней новый лист
        ws.title = "Sheet1"  # Называем его

        ws.append(reader.fieldnames)  # Добавляем заголовок
        for row in csv_file:
            ws.append([row[field] for field in reader.fieldnames])  # В каждую строку таблицы добавляем соответствующую ей информацию
        
        for column in ws.columns:   # Равняем всё по ширине, 8 - минимум
            max_length = 8
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                max_length = max(len(str(cell.value)), max_length)

            ws.column_dimensions[column_letter].width = max_length
            

    wb.save(xlsx_path)  # Сохраняем все изменения в указанном пути